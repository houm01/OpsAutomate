#!/usr/bin/env python3

import os
import shutil
import yaml

from typing import Literal
from pathlib import Path
from openpyxl import load_workbook
from nb_log import get_logger
from jinja2 import FileSystemLoader, Environment



log = get_logger('build_telegraf_config')

config_path = Path.cwd() / 'telegraf_config'
# log.debug(config_path)
jinja2_path = config_path / 'jinja2'
output_path = config_path / 'output'
config_yaml = config_path / 'edit' / 'config.yaml'
resources_yaml = config_path / 'edit' / 'resources.yaml'

shutil.rmtree(output_path)
os.makedirs(output_path)
log.warning('删除 output 目录并重新创建..')

def convert_dict(original_dict):
    # 初始化目标字典
    result_dict = {}

    # 遍历原始字典
    for ip, details in original_dict.items():
        source = details['source']
        
        # 确保目标字典中存在源键
        if source not in result_dict:
            result_dict[source] = {}
        
        # 将 IP 地址和详细信息添加到目标字典中
        result_dict[source][ip] = details

    return result_dict




def get_resources_by_yaml():
    with open(resources_yaml, 'r', encoding='utf-8') as f:
        return yaml.safe_load(f.read())


class BuildTelegrafConfig:
    '''
    _summary_
    '''
    def __init__(self,
                 init_type: Literal['xlsx', 'yaml'],
                 config_yaml: str=None,
                 resource_yaml: str=None,
                 resource_xlsx: str=None,
                 resource_path: str='./',
                 output_path: str='./'):

        self.init_type = init_type
        
        self.config_yaml = config_yaml
        self.resource_yaml = resource_yaml
        self.resource_xlsx = resource_xlsx
        self.resource_path = resource_path
        self.output_path = output_path
        
    def init_config(self):
        '''
        生成示例配置, 指导用户进行下一步操作

        Args:
            init_type (Literal[xlsx, yaml]): _description_
        '''
        if self.init_type == 'xlsx':
            
            if not os.path.exists(self.resource_path + '/sample_input.xlsx'):
                os.makedirs(self.resource_path)
                shutil.copy(src=Path.cwd() / 'telegraf_config' / 'example' / 'sample.xlsx', dst=self.resource_path + '/sample_input.xlsx')
            if not os.path.exists(self.resource_path + '/config_input.yaml'):
                # os.makedirs(self.resource_path)
                shutil.copy(src=Path.cwd() / 'telegraf_config' / 'edit' / 'config.yaml', dst=self.resource_path + '/config_input.yaml')
    
    def get_config(self):
        with open(self.resource_path + '/config_input.yaml', 'r', encoding='utf-8') as f:
            return yaml.safe_load(f.read())

    def render_config(self, location, monitor_type, config, instances):
        
        j2_loader = FileSystemLoader(jinja2_path)
        env = Environment(loader=j2_loader)
        j2_tmpl = env.get_template('telegraf_config.j2')

        result = j2_tmpl.render(config=config, monitor_type=monitor_type, instances=instances)
        
        log.debug(f'output_path: [{self.output_path}]')
        if not os.path.exists(self.output_path):
            os.makedirs(self.output_path) 
        with open(self.output_path + f'telegraf_{location}_{monitor_type}.conf', 'w') as f:
            f.write(result)


        
    def output(self):
        if self.init_type == 'xlsx':
            # 加载工作簿
            workbook = load_workbook(self.resource_path + '/sample_input.xlsx')

            # 选择活动的工作表
            sheet_snmp = workbook['snmp']
            sheet_ping = workbook['ping']
            sheet_http = workbook['http']
            sheet_tcp = workbook['tcp']
            sheet_ssl = workbook['ssl']
            sheet_dns = workbook['dns']
            
            sheet_prometheus = workbook['prometheus']
            
            resources = {}
            instance_snmp = {}
            instance_ping = {}
            instance_http = {}
            instance_ssl = {}
            instance_tcp = {}
            instance_dns = {}
            instance_prometheus = {}
            # 读取所有的数据
            for row in sheet_snmp.iter_rows(values_only=True):
                if row[2] != 'ip':
                    instance_snmp[row[2]] = {
                        'hostname': row[1],
                        'ip': row[2],
                        'port': row[3],
                        'snmp_version': row[5],
                        'vendor': row[4],
                        'snmp_auth': row[6],
                        'source': row[0]
                    }

            for row in sheet_ping.iter_rows(values_only=True):
                if row[1] != 'ip' and row[2] is not None:
                    instance_ping[row[1]] = {
                        'name': row[2],
                        'source': row[0]
                    }   

            for row in sheet_http.iter_rows(values_only=True):
                if row[1] != 'url' and row[2] is not None:
                    instance_http[row[1]] = {
                        'name': row[2],
                        'source': row[0]
                    } 

            for row in sheet_ssl.iter_rows(values_only=True):
                if row[1] != 'url' and row[2] is not None:
                    instance_ssl[row[1]] = {
                        'name': row[2],
                        'source': row[0]
                    } 

            for row in sheet_tcp.iter_rows(values_only=True):
                    if row[1] != 'url' and row[2] is not None:
                        instance_tcp[row[1]] = {
                            'name': row[2],
                            'source': row[0]
                        } 

            for row in sheet_dns.iter_rows(values_only=True):
                    if row[1] != 'server' and row[2] is not None:
                        instance_dns[row[1]] = {
                            'domain': row[2],
                            'record_type': row[3],
                            'source': row[0]
                        } 


            for row in sheet_prometheus.iter_rows(values_only=True):
                if row[1] != 'url' and row[2] is not None:
                    instance_prometheus[row[1]] = {
                        'name': row[2],
                        'source': row[0]
                    } 
            
            resources['snmp'] = convert_dict(original_dict=instance_snmp)
            resources['ping'] = convert_dict(original_dict=instance_ping)
            resources['http'] = convert_dict(original_dict=instance_http)
            resources['tcp'] = convert_dict(original_dict=instance_tcp)
            resources['ssl'] = convert_dict(original_dict=instance_ssl)
            resources['dns'] = convert_dict(original_dict=instance_dns)
            resources['prometheus'] = convert_dict(original_dict=instance_prometheus)
            
            
            for monitor_type, values in resources.items():
                for location, instances in values.items():
                    self.render_config(location=location, monitor_type=monitor_type, config=self.get_config(), instances=instances)
        
        elif self.init_type == 'yaml':
            for monitor_type, values in get_resources_by_yaml().items():
                for location, instances in values.items():
                    # log.warning(location)
                    self.render_config(location=location, monitor_type=monitor_type, config=self.get_config(), instances=instances)


